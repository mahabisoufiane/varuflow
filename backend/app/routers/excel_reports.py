"""Excel / XLSX report generation.

Endpoints:
  GET /api/reports/excel/invoices     → invoices with subtotal/VAT/total formulas
  GET /api/reports/excel/inventory    → stock levels with stock-value formula
  GET /api/reports/excel/customers    → customer list with invoiced/paid balance
  GET /api/reports/excel/revenue      → monthly revenue analytics
"""
from __future__ import annotations

import io
import logging
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

from app.database import get_db
from app.middleware.auth import get_current_member
from app.models.invoicing import Invoice, InvoiceStatus, Customer, Payment
from app.models.inventory import Product, StockLevel

logger = logging.getLogger(__name__)
router = APIRouter(tags=["excel_reports"])

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_VARUFLOW_BLUE = "1E40AF"
_LIGHT_GRAY = "F8FAFC"
_WHITE = "FFFFFF"
_BORDER_COLOR = "CBD5E1"


def _check_openpyxl() -> None:
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Excel export requires openpyxl. Install it: pip install openpyxl",
        )


def _thin_side() -> "Side":
    return Side(border_style="thin", color=_BORDER_COLOR)


def _thin_border() -> "Border":
    s = _thin_side()
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(ws, row: int, num_cols: int) -> None:
    fill = PatternFill("solid", fgColor=_VARUFLOW_BLUE)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    border = _thin_border()
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_row(ws, row: int, num_cols: int) -> None:
    fill = PatternFill("solid", fgColor=_LIGHT_GRAY if row % 2 == 0 else _WHITE)
    border = _thin_border()
    font = Font(name="Calibri", size=10)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = border
        cell.font = font


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _workbook_to_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_CONTENT_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──────────────────────────────────────────────────────────────────────────────
# Invoices report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/api/reports/excel/invoices")
async def export_invoices_excel(
    from_date: Optional[date] = Query(None, alias="from"),
    to_date: Optional[date] = Query(None, alias="to"),
    status: Optional[str] = Query(None),
    member=Depends(get_current_member),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Download invoices as XLSX with subtotal/VAT/total formula columns."""
    _check_openpyxl()
    try:
        org_id = member["org_id"]

        filters = [Invoice.org_id == org_id]
        if from_date:
            filters.append(Invoice.issue_date >= from_date)
        if to_date:
            filters.append(Invoice.issue_date <= to_date)
        if status:
            filters.append(Invoice.status == status)

        # Fetch invoices with customer name via join
        rows = (await db.execute(
            select(Invoice, Customer.company_name)
            .join(Customer, Invoice.customer_id == Customer.id)
            .where(and_(*filters))
            .order_by(Invoice.issue_date.desc())
            .limit(5000)
        )).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.freeze_panes = "A2"

        headers = [
            "Invoice #", "Customer", "Issue Date", "Due Date", "Status",
            "Subtotal (excl. VAT)", "VAT Amount", "Total (incl. VAT)", "Currency", "Notes",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        _style_header_row(ws, 1, len(headers))

        for r_idx, (inv, customer_name) in enumerate(rows, start=2):
            ws.cell(row=r_idx, column=1, value=inv.invoice_number or "")
            ws.cell(row=r_idx, column=2, value=customer_name or "")
            ws.cell(row=r_idx, column=3, value=inv.issue_date)
            ws.cell(row=r_idx, column=4, value=inv.due_date)
            ws.cell(row=r_idx, column=5, value=str(inv.status.value if hasattr(inv.status, "value") else inv.status))
            sub_cell = ws.cell(row=r_idx, column=6, value=float(inv.subtotal or 0))
            sub_cell.number_format = '#,##0.00'
            vat_cell = ws.cell(row=r_idx, column=7, value=float(inv.vat_amount or 0))
            vat_cell.number_format = '#,##0.00'
            # Excel formula: subtotal + VAT
            total_cell = ws.cell(row=r_idx, column=8, value=f"=F{r_idx}+G{r_idx}")
            total_cell.number_format = '#,##0.00'
            ws.cell(row=r_idx, column=9, value=inv.currency or "SEK")
            ws.cell(row=r_idx, column=10, value=inv.notes or "")
            _style_data_row(ws, r_idx, len(headers))

        if rows:
            last_data = len(rows) + 1
            total_row = last_data + 1
            ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True, name="Calibri")
            for col, lbl in [(6, "F"), (7, "G"), (8, "H")]:
                c = ws.cell(row=total_row, column=col, value=f"=SUM({lbl}2:{lbl}{last_data})")
                c.number_format = '#,##0.00'
                c.font = Font(bold=True, name="Calibri")

        _set_col_widths(ws, [14, 28, 12, 12, 12, 20, 16, 20, 10, 30])

        date_str = date.today().strftime("%Y%m%d")
        return _workbook_to_response(wb, f"invoices_{date_str}.xlsx")

    except HTTPException:
        raise
    except Exception as e:
        logger.error("export_invoices_excel failed: %s", e, extra={"org_id": str(org_id)})
        raise HTTPException(status_code=500, detail="Internal server error")


# ──────────────────────────────────────────────────────────────────────────────
# Inventory / stock report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/api/reports/excel/inventory")
async def export_inventory_excel(
    low_stock_only: bool = Query(False),
    member=Depends(get_current_member),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Download inventory stock levels as XLSX with stock-value formula."""
    _check_openpyxl()
    try:
        org_id = member["org_id"]

        products = (await db.execute(
            select(Product)
            .where(and_(Product.org_id == org_id, Product.is_active == True))
            .order_by(Product.name)
            .limit(10000)
        )).scalars().all()

        stock_map: dict[uuid.UUID, StockLevel] = {}
        if products:
            sl_rows = (await db.execute(
                select(StockLevel).where(and_(
                    StockLevel.org_id == org_id,
                    StockLevel.product_id.in_([p.id for p in products]),
                ))
            )).scalars().all()
            for sl in sl_rows:
                # Keep the highest-quantity entry per product (multiple warehouses)
                existing = stock_map.get(sl.product_id)
                if existing is None:
                    stock_map[sl.product_id] = sl
                else:
                    stock_map[sl.product_id].quantity += sl.quantity  # type: ignore[assignment]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.freeze_panes = "A2"

        headers = [
            "SKU", "Product Name", "Category", "Unit", "Purchase Price",
            "Sale Price", "Qty On Hand", "Reorder Level", "Stock Value (Cost)",
            "Status",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        _style_header_row(ws, 1, len(headers))

        r_idx = 2
        for prod in products:
            sl = stock_map.get(prod.id)
            qty = float(sl.quantity) if sl else 0.0
            reorder = float(prod.reorder_level or 0)

            if low_stock_only and qty > reorder:
                continue

            status = "OK"
            if qty <= 0:
                status = "OUT OF STOCK"
            elif qty <= reorder:
                status = "LOW STOCK"

            ws.cell(row=r_idx, column=1, value=prod.sku or "")
            ws.cell(row=r_idx, column=2, value=prod.name or "")
            ws.cell(row=r_idx, column=3, value=prod.category or "")
            ws.cell(row=r_idx, column=4, value=prod.unit or "st")
            cost_cell = ws.cell(row=r_idx, column=5, value=float(prod.purchase_price or 0))
            cost_cell.number_format = '#,##0.00'
            price_cell = ws.cell(row=r_idx, column=6, value=float(prod.sell_price or 0))
            price_cell.number_format = '#,##0.00'
            ws.cell(row=r_idx, column=7, value=qty)
            ws.cell(row=r_idx, column=8, value=reorder)
            # Stock value = purchase_price * qty (formula)
            value_cell = ws.cell(row=r_idx, column=9, value=f"=E{r_idx}*G{r_idx}")
            value_cell.number_format = '#,##0.00'
            ws.cell(row=r_idx, column=10, value=status)
            _style_data_row(ws, r_idx, len(headers))
            r_idx += 1

        if r_idx > 2:
            last_data = r_idx - 1
            total_row = r_idx
            ws.cell(row=total_row, column=8, value="TOTAL STOCK VALUE").font = Font(bold=True, name="Calibri")
            c = ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{last_data})")
            c.number_format = '#,##0.00'
            c.font = Font(bold=True, name="Calibri")

        _set_col_widths(ws, [14, 32, 16, 8, 16, 14, 12, 14, 20, 14])

        date_str = date.today().strftime("%Y%m%d")
        return _workbook_to_response(wb, f"inventory_{date_str}.xlsx")

    except HTTPException:
        raise
    except Exception as e:
        logger.error("export_inventory_excel failed: %s", e, extra={"org_id": str(org_id)})
        raise HTTPException(status_code=500, detail="Internal server error")


# ──────────────────────────────────────────────────────────────────────────────
# Customer statement report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/api/reports/excel/customers")
async def export_customers_excel(
    member=Depends(get_current_member),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Download customer list with outstanding balance as XLSX."""
    _check_openpyxl()
    try:
        org_id = member["org_id"]

        customers = (await db.execute(
            select(Customer)
            .where(and_(Customer.org_id == org_id, Customer.is_active == True))
            .order_by(Customer.company_name)
            .limit(5000)
        )).scalars().all()

        # Aggregate invoiced and paid per customer (SENT + PAID only — exclude drafts)
        invoiced_rows = (await db.execute(
            select(
                Invoice.customer_id,
                func.coalesce(func.sum(Invoice.total_sek), 0).label("total_invoiced"),
            )
            .where(and_(
                Invoice.org_id == org_id,
                Invoice.status.in_([InvoiceStatus.SENT, InvoiceStatus.PAID]),
            ))
            .group_by(Invoice.customer_id)
        )).all()
        invoiced_map = {str(r.customer_id): float(r.total_invoiced) for r in invoiced_rows}

        paid_rows = (await db.execute(
            select(
                Invoice.customer_id,
                func.coalesce(func.sum(Payment.amount), 0).label("total_paid"),
            )
            .join(Payment, Payment.invoice_id == Invoice.id)
            .where(and_(Invoice.org_id == org_id))
            .group_by(Invoice.customer_id)
        )).all()
        paid_map = {str(r.customer_id): float(r.total_paid) for r in paid_rows}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.freeze_panes = "A2"

        headers = [
            "Customer Name", "Email", "Phone", "VAT Number",
            "Total Invoiced", "Total Paid", "Outstanding Balance",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        _style_header_row(ws, 1, len(headers))

        for r_idx, cust in enumerate(customers, start=2):
            invoiced = invoiced_map.get(str(cust.id), 0.0)
            paid = paid_map.get(str(cust.id), 0.0)

            ws.cell(row=r_idx, column=1, value=cust.company_name or "")
            ws.cell(row=r_idx, column=2, value=cust.email or "")
            ws.cell(row=r_idx, column=3, value=cust.phone or "")
            ws.cell(row=r_idx, column=4, value=cust.vat_number or "")
            inv_cell = ws.cell(row=r_idx, column=5, value=invoiced)
            inv_cell.number_format = '#,##0.00'
            paid_cell = ws.cell(row=r_idx, column=6, value=paid)
            paid_cell.number_format = '#,##0.00'
            # Outstanding = invoiced - paid (formula)
            balance_cell = ws.cell(row=r_idx, column=7, value=f"=E{r_idx}-F{r_idx}")
            balance_cell.number_format = '#,##0.00'
            _style_data_row(ws, r_idx, len(headers))

        if customers:
            last = len(customers) + 1
            total_row = last + 1
            ws.cell(row=total_row, column=4, value="TOTALS").font = Font(bold=True, name="Calibri")
            for col, lbl in [(5, "E"), (6, "F"), (7, "G")]:
                c = ws.cell(row=total_row, column=col, value=f"=SUM({lbl}2:{lbl}{last})")
                c.number_format = '#,##0.00'
                c.font = Font(bold=True, name="Calibri")

        _set_col_widths(ws, [30, 28, 16, 16, 18, 14, 20])

        date_str = date.today().strftime("%Y%m%d")
        return _workbook_to_response(wb, f"customers_{date_str}.xlsx")

    except HTTPException:
        raise
    except Exception as e:
        logger.error("export_customers_excel failed: %s", e, extra={"org_id": str(org_id)})
        raise HTTPException(status_code=500, detail="Internal server error")


# ──────────────────────────────────────────────────────────────────────────────
# Revenue analytics report
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/api/reports/excel/revenue")
async def export_revenue_excel(
    year: int = Query(default=0),
    member=Depends(get_current_member),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Download monthly revenue breakdown as XLSX with formula totals."""
    _check_openpyxl()
    try:
        org_id = member["org_id"]
        if not year:
            year = date.today().year

        rows = (await db.execute(
            select(
                func.extract("month", Invoice.issue_date).label("month"),
                func.coalesce(func.sum(Invoice.subtotal), 0).label("subtotal"),
                func.coalesce(func.sum(Invoice.vat_amount), 0).label("vat"),
                func.coalesce(func.sum(Invoice.total_sek), 0).label("total"),
                func.count(Invoice.id).label("count"),
            )
            .where(and_(
                Invoice.org_id == org_id,
                Invoice.status.in_([InvoiceStatus.SENT, InvoiceStatus.PAID]),
                func.extract("year", Invoice.issue_date) == year,
            ))
            .group_by(func.extract("month", Invoice.issue_date))
            .order_by(func.extract("month", Invoice.issue_date))
        )).all()

        month_data = {int(r.month): r for r in rows}
        month_names = [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Revenue {year}"
        ws.freeze_panes = "A2"

        headers = ["Month", "# Invoices", "Subtotal (excl. VAT)", "VAT", "Total Revenue", "Cumulative Total"]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        _style_header_row(ws, 1, len(headers))

        for m in range(1, 13):
            r_idx = m + 1
            r = month_data.get(m)

            ws.cell(row=r_idx, column=1, value=month_names[m - 1])
            ws.cell(row=r_idx, column=2, value=int(r.count) if r else 0)
            sub_cell = ws.cell(row=r_idx, column=3, value=float(r.subtotal) if r else 0.0)
            sub_cell.number_format = '#,##0.00'
            vat_cell = ws.cell(row=r_idx, column=4, value=float(r.vat) if r else 0.0)
            vat_cell.number_format = '#,##0.00'
            # Total = subtotal + VAT
            total_cell = ws.cell(row=r_idx, column=5, value=f"=C{r_idx}+D{r_idx}")
            total_cell.number_format = '#,##0.00'
            # Cumulative = SUM E2 to current row
            cum_cell = ws.cell(row=r_idx, column=6, value=f"=SUM($E$2:E{r_idx})")
            cum_cell.number_format = '#,##0.00'
            _style_data_row(ws, r_idx, len(headers))

        # Annual totals
        total_row = 14
        ws.cell(row=total_row, column=1, value=f"TOTAL {year}").font = Font(bold=True, name="Calibri")
        ws.cell(row=total_row, column=2, value="=SUM(B2:B13)").font = Font(bold=True, name="Calibri")
        for col, lbl in [(3, "C"), (4, "D"), (5, "E")]:
            c = ws.cell(row=total_row, column=col, value=f"=SUM({lbl}2:{lbl}13)")
            c.number_format = '#,##0.00'
            c.font = Font(bold=True, name="Calibri")

        _set_col_widths(ws, [14, 12, 22, 16, 18, 18])

        return _workbook_to_response(wb, f"revenue_{year}.xlsx")

    except HTTPException:
        raise
    except Exception as e:
        logger.error("export_revenue_excel failed: %s", e, extra={"org_id": str(org_id)})
        raise HTTPException(status_code=500, detail="Internal server error")
